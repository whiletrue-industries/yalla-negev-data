from google.cloud import firestore
import google.auth
import json
from datetime import datetime
import openpyxl
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.auth.transport.requests import Request
import os

def flatten_document(doc_snapshot):
    """
    Flatten the document to a single-level dictionary, adding document id to each dict.
    Handles nested dictionaries by flattening them to a single level.
    """
    def flatten(nested_dict, parent_key='', sep='.'):
        """
        Recursively flattens a nested dictionary.
        """
        items = []
        for k, v in nested_dict.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(flatten(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
        return dict(items)

    flattened_doc = flatten(doc_snapshot.to_dict())
    flattened_doc['id'] = doc_snapshot.id
    return flattened_doc

def default_serializer(obj):
    """
    Custom serializer function to handle non-serializable types such as Firestore timestamp or DocumentReference.
    """
    if isinstance(obj, datetime):
        return obj.isoformat()
    elif isinstance(obj, firestore.DocumentReference):
        return obj.path
    raise TypeError(f"Type {type(obj)} not serializable")

def read_firestore_subcollections(document_path):
    """
    Reads all sub-collections of a specific document into a dictionary of lists of flattened documents.
    """
    db = firestore.Client()
    document_ref = db.document(document_path)
    subcollections = document_ref.collections()

    all_data = {}
    for subcollection in subcollections:
        subcollection_name = subcollection.id
        docs = subcollection.stream()
        all_data[subcollection_name] = [flatten_document(doc) for doc in docs]
    
    return all_data

# def write_firestore_subcollections(document_path, data_dict):
#     """
#     Writes the list of documents back to Firestore sub-collections, overwriting existing data.
#     """
#     db = firestore.Client()
#     for subcollection_name, data_list in data_dict.items():
#         subcollection_ref = db.document(document_path).collection(subcollection_name)
        
#         for doc in data_list:
#             # Assumes each object in list has an 'id' field, which is used as document id
#             if 'id' in doc:
#                 doc_ref = subcollection_ref.document(doc['id'])
#                 # Remove 'id' before updating Firestore
#                 data_to_write = {k: v for k, v in doc.items() if k != 'id'}
#                 doc_ref.set(data_to_write)
#             else:
#                 # If 'id' is not present, add new document without specifying document id
#                 subcollection_ref.add(doc)

def process_surveys(surveys):
    """
    Process the surveys data to create a new list of surveys with additional fields.
    """
    processed_surveys = {}
    for survey in surveys:
        name = survey.get('name.he', survey.get('name.en', ''))
        description = survey.get('description.he', survey.get('description.en', ''))
        if not name:
            continue
        questions = []
        for question in survey.get('questions', []):
            text = question.get('text', {})
            question_text = text.get('he', text.get('en', ''))
            questions.append({
                'id': question.get('id'),
                'text': question_text,
            })
        if not questions:
            continue
        print(f"Processing survey: {name} ({description}), with {len(questions)} questions")
        processed_survey = {
            'name': name,
            'description': description,
            'created_at': survey['creationDateTime'].isoformat(),
            'questions': questions,
        }
        processed_surveys[survey['id']] = processed_survey
    return processed_surveys

def process_responses(responses, survey_id, survey):
    """
    Process the responses data to create a new list of responses with additional fields.
    """
    processed_responses = []
    headers = ['time', 'lat', 'lon'] + [q['text'] for q in survey['questions']]
    for response in responses:
        if response['surveyId'] != survey_id:
            continue
        if 'coordinates.latitude' not in response:
            print(f'Response missing coordinate data: {response["id"]}')
            continue
        answers = dict(
            time=response['submittedTs'].isoformat(),
            lat=str(response['coordinates.latitude']),
            lon=str(response['coordinates.longitude']),
        )
        response_answers = response.get('responses', [])
        for question in survey['questions']:
            answer = [a for a in response_answers if a['questionId'] == question['id']]
            if len(answer) == 1:
                answers[question['text']] = str(answer[0]['response'])
        processed_responses.append(answers)
    return headers, processed_responses

def fix_sheet(sheet):
    print('Fixing sheet...', sheet.title)
    # Auto-size columns
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            cell.alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center", wrap_text=False, readingOrder=2)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.0
        current_width = sheet.column_dimensions[column[0].column_letter].width
        # print(f"Column {column[0].column_letter} width: {current_width} -> {adjusted_width}")
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    sheet.sheet_view.rightToLeft = True

def write_to_excel(surveys, sheets):
    # Create a new workbook
    wb = openpyxl.Workbook()

    # Remove the default sheet created by openpyxl
    default_sheet = wb.active
    wb.remove(default_sheet)

    surveys_sheet = wb.create_sheet(title="סקרים")
    surveys_sheet.cell(row=1, column=1, value="שם")
    surveys_sheet.cell(row=1, column=2, value="תיאור")
    surveys_sheet.cell(row=1, column=3, value="נוצר ב")
    surveys_sheet.cell(row=1, column=4, value="מספר שאלות")
    surveys_sheet.cell(row=1, column=5, value="מספר תגובות")
    for i, survey in enumerate(surveys.values(), start=2):
        surveys_sheet.cell(row=i, column=1, value=survey['name'])
        surveys_sheet.cell(row=i, column=2, value=survey['description'])
        surveys_sheet.cell(row=i, column=3, value=survey['created_at'])
        surveys_sheet.cell(row=i, column=4, value=len(survey['questions']))
        responses = [r for s, _, r in sheets if s == survey['name']]
        surveys_sheet.cell(row=i, column=5, value=len(responses))
    fix_sheet(surveys_sheet)

    # Create each sheet from the data dictionary
    for sheet_name, headers, responses in sheets:
        # Create a new sheet with the specified name
        sheet = wb.create_sheet(title=sheet_name)

        # Write headers to the first row
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_index, value=header)

        # Write response data starting from the second row
        for row_index, response in enumerate(responses, start=2):
            for col_index, header in enumerate(headers, start=1):
                sheet.cell(row=row_index, column=col_index, value=response.get(header, ""))
        fix_sheet(sheet)

    # Save the workbook to a file
    current_date = datetime.now().strftime("%Y-%m-%d")
    output_filename = f'yallanegev-{current_date}.xlsx'
    wb.save(output_filename)

    # Upload the output file to Google Drive, using credentials in GOOGLE_APPLICATION_CREDENTIALS, using google-api-python-client
    DRIVE_FOLDER = os.getenv('DRIVE_FOLDER_ID').strip()  # Use environment variable for Drive folder ID

    # Authenticate and create the Drive API client
    creds = None
    credentials_path = os.getenv('GOOGLE_APPLICATION_CREDENTIALS')
    if credentials_path and os.path.exists(credentials_path):
        creds = google.auth.load_credentials_from_file(credentials_path)[0]
    else:
        raise Exception("Credentials file not found or GOOGLE_APPLICATION_CREDENTIALS environment variable is not set.")

    service = build('drive', 'v3', credentials=creds)

    # Upload the file
    file_metadata = {
        'name': output_filename,
        'parents': [DRIVE_FOLDER.split('/')[-1]]
    }
    media = MediaFileUpload(output_filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    print(f"File ID: {file.get('id')}")

def main():
    document_path = 'versions/v1'  # Replace with your document path

    # Read Firestore Sub-collections
    data_dict = read_firestore_subcollections(document_path)

    # Get Surveys:
    surveys = process_surveys(data_dict.get('surveys', []))

    # Get responses:
    sheets = []
    for survey_id, survey in surveys.items():
        headers, responses = process_responses(data_dict.get('responses', []), survey_id, survey)
        print(f"Survey: {survey['name']} ({survey['description']})")
        print(f"Headers: {headers}")
        print(f"# Responses: {len(responses)}")
        if responses:
            print(f"Responses: {responses[0]}")
            sheets.append((survey['name'], headers, responses))

    # Write to Excel
    write_to_excel(surveys, sheets)
if __name__ == '__main__':
    main()
